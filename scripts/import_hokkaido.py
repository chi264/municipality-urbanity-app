import json
import numbers
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "sources" / "hokkaido_municipalities_2025.xlsx"
MUNICIPALITIES_PATH = ROOT / "data" / "municipalities.json"
TRAVEL_PATH = ROOT / "data" / "travel_times.json"

WIDE_REGION = {
    "石狩振興局": "道央",
    "空知総合振興局": "道央",
    "後志総合振興局": "道央",
    "胆振総合振興局": "道央",
    "日高振興局": "道央",
    "渡島総合振興局": "道南",
    "檜山振興局": "道南",
    "上川総合振興局": "道北",
    "留萌振興局": "道北",
    "宗谷総合振興局": "道北",
    "オホーツク総合振興局": "道東",
    "十勝総合振興局": "道東",
    "釧路総合振興局": "道東",
    "根室振興局": "道東",
}

CENTER_SUFFIX = {
    "市": "役所周辺",
    "町": "役場周辺",
    "村": "役場周辺",
}

OLD_TRAVEL_IDS = {
    "hokkaido-sapporo": "hokkaido-011002",
    "hokkaido-asahikawa": "hokkaido-012041",
    "hokkaido-hakodate": "hokkaido-012025",
    "hokkaido-otaru": "hokkaido-012033",
    "hokkaido-niseko": "hokkaido-013951",
}


def build_hokkaido_rows():
    wb = openpyxl.load_workbook(SOURCE, data_only=True)
    ws = wb.active
    current_subprefecture = None
    rows = []

    for row_index in range(25, ws.max_row + 1):
        name = ws.cell(row_index, 1).value
        code = ws.cell(row_index, 2).value
        if not name:
            continue

        name = str(name).strip().replace("\n", "")
        if code is None:
            if name.endswith("局"):
                current_subprefecture = name
            continue

        population = ws.cell(row_index, 9).value
        if name.endswith("区"):
            continue
        if not isinstance(population, numbers.Number):
            continue

        code = str(code).strip()
        area = ws.cell(row_index, 3).value
        density = ws.cell(row_index, 17).value
        municipality_type = name[-1]
        subprefecture = current_subprefecture.replace("\n", "") if current_subprefecture else None

        rows.append(
            {
                "id": f"hokkaido-{code}",
                "localGovernmentCode": code,
                "prefecture": "北海道",
                "name": name,
                "type": municipality_type,
                "population": int(population),
                "area": round(float(area), 2) if isinstance(area, numbers.Number) else None,
                "populationDensity": round(float(density), 2) if isinstance(density, numbers.Number) else None,
                "region": WIDE_REGION.get(subprefecture, subprefecture),
                "subprefecture": subprefecture,
                "center": {
                    "name": f"{name}{CENTER_SUFFIX.get(municipality_type, '役場周辺')}",
                },
            }
        )

    return rows


def main():
    hokkaido_rows = build_hokkaido_rows()
    municipalities = json.loads(MUNICIPALITIES_PATH.read_text(encoding="utf-8"))
    municipalities = [
        m
        for m in municipalities
        if m.get("prefecture") != "北海道" and not str(m.get("id", "")).startswith("hokkaido-")
    ]
    municipalities.extend(hokkaido_rows)
    MUNICIPALITIES_PATH.write_text(
        json.dumps(municipalities, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    travel_times = json.loads(TRAVEL_PATH.read_text(encoding="utf-8"))
    for item in travel_times:
        old_id = item.get("toMunicipalityId")
        if old_id in OLD_TRAVEL_IDS:
            item["toMunicipalityId"] = OLD_TRAVEL_IDS[old_id]
    TRAVEL_PATH.write_text(
        json.dumps(travel_times, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    print(f"Hokkaido municipalities: {len(hokkaido_rows)}")
    print(f"Total municipalities: {len(municipalities)}")


if __name__ == "__main__":
    main()
